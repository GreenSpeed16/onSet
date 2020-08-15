from tkinter import *
import tkinter.messagebox
import openpyxl

del_root = Tk()
delete_list = {}
def deleteRoute():
    global delete_list
    def confirmDelete():
        global delete_list
        deleted = 0
        for key, value in delete_list.items():
            if value:
                deleted += 1

        # Confirm deletion
        confirm_answer = tkinter.messagebox.askquestion('Confirm Deletion', 'Your are about to delete {} '
                                                                                   'Routes. Continue?'.format(deleted))

        if confirm_answer == 'yes':
            for key2, value2 in delete_list.items():
                if value:
                    boulder_sheet.delete_rows(key2)
                    delete_list = {key2 + 1: value2}
            tkinter.messagebox.showinfo('', 'Deleted {} routes successfully.'.format(deleted))
            delete_window.destroy()
        else:
            tkinter.messagebox.showinfo('Successful Cancel', 'Canceled route deletion.')
        route_book.save('Routes.xlsx')

    # Change buttons to represent excluded routes
    def btnPress(num, row):
        if button_list[num-1].cget('fg') == 'red':
            delete_list[row] = False
            button_list[num-1].configure(fg='green')
        else:
            delete_list[row] = True
            button_list[num-1].configure(fg='red')

    route_book = openpyxl.load_workbook('Routes.xlsx')
    boulder_sheet = route_book['Boulders']

    delete_window = Tk()

    button_list = []

    num_buttons = 0

    info_label = Label(delete_window, text='Pressing confirm will delete all routes highlighted red')
    info_label.grid(row=0, column=1)

    for cell in range(1, boulder_sheet.max_row + 1):
        grade = boulder_sheet.cell(row=cell, column=1).value
        wall = boulder_sheet.cell(row=cell, column=2).value
        setter = boulder_sheet.cell(row=cell, column=3).value
        color = boulder_sheet.cell(row=cell, column=4).value

        if wall == wallDrop.get():
            num_buttons += 1
            button_list.append(Button(delete_window, fg='red', text='{}, {}, {}, {}'.format(color, grade, wall, setter),
                                     command=lambda x=num_buttons, y = cell: btnPress(x, y)))
            button_list[num_buttons - 1].grid(row=num_buttons, column=1)
            delete_list.update({cell: True})

    confirm_button = Button(delete_window, text='Confirm Deletion', command=confirmDelete)
    confirm_button.grid(row=num_buttons+1, column=1)

wall_options = ['Wall:', 'Topout', 'Main']
setter_options = ['Setter:', 'Chase', 'Chris', 'Christian', 'Jeff', 'Jeremy', 'Joey', 'Mitch']
color_options = ['Color:', 'Green', 'Orange', 'Pink']

wallDrop = StringVar()
wallDrop.set(wall_options[0])

setterDrop = StringVar()
setterDrop.set(setter_options[0])

colorDrop = StringVar()
colorDrop.set(color_options[0])

wallDelMenu = OptionMenu(del_root, wallDrop, *wall_options)
wallDelMenu.grid(row=0, column=1)

setterDelMenu = OptionMenu(del_root, setterDrop, *setter_options)
setterDelMenu.grid(row=0, column=2)

colorDelMenu = OptionMenu(del_root, colorDrop, *color_options)
colorDelMenu.grid(row=0, column=3)

submitButton = Button(del_root, text='Search For Routes', command=deleteRoute)
submitButton.grid(row=1, column=1, columnspan=2)

del_root.mainloop()
