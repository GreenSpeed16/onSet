# Import openpyxl
import openpyxl

# Open the excel document
routeBook = openpyxl.load_workbook('Routes.xlsx')
boulderSheet = routeBook['Boulders']
ropeSheet = routeBook['Ropes']

# Function to ask the user for the route
def askRoute():
    routeOrBoulder = str(input('Are you entering rope routes or boulder problems? '))

    if routeOrBoulder.lower() == 'boulder' or routeOrBoulder.lower() == 'boulders':
        newRoute = Boulder(int(input('What grade are the problems? V')),
                           str(input('What wall are the problems set on? ')),
                           int(input('How many of this grade were set? ')))
    elif routeOrBoulder.lower() == 'rope' or routeOrBoulder.lower() == 'route' or routeOrBoulder == 'ropes':
        newRoute = Ropes(str(input('What grade are the routes? 5.')),
                           str(input('What wall were the problems set on? ')),
                           int(input('How many routes of this grade did you set? ')))
    else:
        print('You did not enter any of the acceptable options: boulder, boulder problem, rope, route, or rope route'
            + '. (Not case sensitive.)')
        askRoute()

# Create a ropes class
class Ropes():
    def __init__(self, grade, wall, amount):
        # Convert all strings to lowercase
        grade = grade.lower()
        wall = wall.lower()

        # Update spreadsheet
        for i in range(0,amount):
            ropeSheet.cell(column=1, row=ropeSheet.max_row + 1).value = '5.' + grade
            ropeSheet.cell(column=2, row=ropeSheet.max_row).value = wall
            ropeSheet.cell(column=3, row=ropeSheet.max_row).value = 'multi-add'

# Create a boulder class
class Boulder():
    def __init__(self, grade, wall, amount):
        # Convert all strings to lowercase and int to string
        grade = str(grade)
        wall = wall.lower()

        # When the object is created, update spreadsheet using parameters
        for i in range(0,amount):
            boulderSheet.cell(column=1, row=boulderSheet.max_row + 1).value = 'v' + grade
            boulderSheet.cell(column=2, row=boulderSheet.max_row).value = wall
            boulderSheet.cell(column=3, row=boulderSheet.max_row).value = 'multi-add'

askRoute()

routeBook.save('Routes.xlsx')