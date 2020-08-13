# Import openpyxl
import openpyxl

# Open the excel document
routeBook = openpyxl.load_workbook('Routes.xlsx')
boulderSheet = routeBook['Boulders']
ropeSheet = routeBook['Ropes']

# Create a list and variables to find deletable routes
delete_list = []
delete_param = 0
find_value = ''

# Ask the user what they are searching for
ask_delete = str(input('Are you searching for a route by grade, wall, or setter? '))
if ask_delete.lower() == 'grade':
    delete_param = 1
elif ask_delete.lower() == 'wall':
    delete_param = 2
elif ask_delete.lower() == 'setter':
    delete_param = 3

# Ask the user what parameter to search by
if delete_param == 1:
    find_value = str(input('What grade do you want to delete? '))
elif delete_param == 2:
    find_value = str(input('What wall do you want to delete? '))
elif delete_param == 3:
    find_value = str(input('What setter do you want to delete? '))

# Iterate over the whole sheet, adding any row with a deletable parameter to the delete list
for cell in range(2, boulderSheet.max_row + 1):
    if boulderSheet.cell(column=delete_param, row=cell).value == str(find_value).lower():
        delete_list.append(cell)

# Confirm deletion
y_or_n = str(input('Found ' + str(len(delete_list)) + ' routes that meet criteria, delete all? y/n: '))

# Iterate over the delete list, deleting the entire row if answer is yes
if y_or_n.lower() == 'y' or y_or_n.lower() == 'yes':
    for row in delete_list:
        boulderSheet.delete_rows(row)
        delete_list = [row - 1 for row in delete_list]
else:
    print('Deletion canceled.')

routeBook.save('Routes.xlsx')