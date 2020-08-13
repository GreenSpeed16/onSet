import matplotlib.pyplot as plt
import openpyxl
import numpy as np


# Get data for current graph
route_book = openpyxl.load_workbook('Routes.xlsx')
boulder_sheet = route_book['Boulders']

grade_list = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

positions = np.arange(len(grade_list))

for cell in range(2, boulder_sheet.max_row + 1):
    temp_grade = boulder_sheet.cell(column=1, row=cell).value
    grade_int = int(temp_grade.replace('V', ''))
    # Fill the list with appropriate values
    if grade_int == 0:
        grade_list[0] += 1
    elif grade_int == 1:
        grade_list[1] += 1
    elif grade_int == 2:
        grade_list[2] += 1
    elif grade_int == 3:
        grade_list[3] += 1
    elif grade_int == 4:
        grade_list[4] += 1
    elif grade_int == 5:
        grade_list[5] += 1
    elif grade_int == 6:
        grade_list[6] += 1
    elif grade_int == 7:
        grade_list[7] += 1
    elif grade_int == 8:
        grade_list[8] += 1
    elif grade_int == 9:
        grade_list[9] += 1

# Get data for goal graph
goal_list = []
data_sheet = route_book['Data']
for cell in range(1, data_sheet.max_row + 1):
    temp_g_grade = data_sheet.cell(row=cell, column=1).value
    goal_list.append(temp_g_grade)

route_book.save('Routes.xlsx')

# Find difference between graphs
grade_difference = []
for i in range(0, 10):
    grade_difference.append(grade_list[i] - goal_list[i] )

x_axis = ['V0',
          'V1',
          'V2',
          'V3',
          'V4',
          'V5',
          'V6',
          'V7',
          'V8',
          'V9']


fig = plt.figure()

plt.bar(positions, grade_list, width=0.5, color='red')
plt.bar(positions+0.5, goal_list, width=0.5)
plt.legend(['Current Spread', 'Goal Spread'])
plt.title('Route Comparison')
plt.xticks(positions+0.25, x_axis)
plt.ylabel('Amount')

plt.show()
