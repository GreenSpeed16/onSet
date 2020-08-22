# Imports
from tkinter import *
import tkinter.messagebox
from tkinter import ttk
import matplotlib.pyplot as plt
import openpyxl
import numpy as np

# Define functions
# Submit route
def submitRoute(sheet):
    # Load workbook
    route_book = openpyxl.load_workbook('Routes.xlsx')
    entry_sheet = route_book[sheet]

    grade_option = gradeDrop.get()
    wall_option = wallDrop.get()
    button = submitButton

    # Detect if a route or boulder is being entered
    if sheet == 'Ropes':
        wall_option = rwallDrop.get()
        grade_option = rGradeDrop.get()
        button = rSubmitButton

    # Update spreadsheet
    entry_sheet.cell(column=1, row=entry_sheet.max_row + 1).value = grade_option
    entry_sheet.cell(column=2, row=entry_sheet.max_row).value = wall_option
    entry_sheet.cell(column=3, row=entry_sheet.max_row).value = setterDrop.get()
    entry_sheet.cell(column=4, row=entry_sheet.max_row).value = colorDrop.get()
    tkinter.messagebox.showinfo(title='Route Added', message='Your route has been successfully added.')
    gradeDrop.set(grade_options[0])
    wallDrop.set(wall_options[0])
    rGradeDrop.set(grade_options[0])
    rwallDrop.set(rwall_options[0])
    setterDrop.set(setter_options[0])
    colorDrop.set(color_options[0])

    button.configure(state=DISABLED)

    # Save workbook
    route_book.save('Routes.xlsx')

# Delete window
def deleteWindow(sheet):
    del_root = tkinter.Toplevel(root)

    global wall_options
    global rwall_options

    def deleteRoute(event):
        global delete_list

        def confirmDelete():
            global delete_list
            deleted = 0
            for key, value in delete_list.items():
                if value:
                    deleted += 1

            # Confirm deletion
            confirm_answer = tkinter.messagebox.askquestion('Confirm Deletion', 'Your are about to delete {} '
                                                                                'Routes. Continue?'.format(deleted))

            if confirm_answer == 'yes':
                for key2, value2 in delete_list.items():
                    if value2:
                        entry_sheet.delete_rows(key2)
                        delete_list = {key2 + 1: value2}
                tkinter.messagebox.showinfo('', 'Deleted {} routes successfully.'.format(deleted))

            else:
                tkinter.messagebox.showinfo('Successful Cancel', 'Canceled route deletion.')
            route_book.save('Routes.xlsx')

            del_root.destroy()
            delete_window.destroy()
            delete_list = {}

        # Change buttons to represent excluded routes
        def btnPress(num, row):
            if button_list[num - 1].cget('fg') == 'red':
                delete_list[row] = False
                button_list[num - 1].configure(fg='green')
            else:
                delete_list[row] = True
                button_list[num - 1].configure(fg='red')

        route_book = openpyxl.load_workbook('Routes.xlsx')
        entry_sheet = route_book[sheet]

        delete_window = tkinter.Toplevel(root)

        button_list = []

        num_buttons = 0

        info_label = Label(delete_window, text='Pressing confirm will delete all routes highlighted red'
                                               '\n click a route to deselect it')
        info_label.grid(row=0, column=1)

        for cell in range(1, entry_sheet.max_row + 1):
            grade = entry_sheet.cell(row=cell, column=1).value
            wall = entry_sheet.cell(row=cell, column=2).value
            setter = entry_sheet.cell(row=cell, column=3).value
            color = entry_sheet.cell(row=cell, column=4).value

            if wall == delWallDrop.get():
                num_buttons += 1
                button_list.append(
                    Button(delete_window, fg='red', text='{}, {}, {}, {}'.format(color, grade, wall, setter),
                           command=lambda x=num_buttons, y=cell: btnPress(x, y)))
                button_list[num_buttons - 1].grid(row=num_buttons, column=1)
                delete_list.update({cell: True})

        del_confirm_button = Button(delete_window, text='Confirm Deletion', command=confirmDelete)
        del_confirm_button.grid(row=num_buttons + 1, column=1)


    if sheet == 'Ropes':
        del_wall_options = rwall_options
    else:
        del_wall_options = wall_options

    delWallDrop = StringVar()
    delWallDrop.set(del_wall_options[0])

    wallDelMenu = OptionMenu(del_root, delWallDrop, *del_wall_options)
    wallDelMenu.grid(row=0, column=1, columnspan=2)

    delSubmitButton = Button(del_root, text='Search For Routes')
    delSubmitButton.bind('<Button-1>', deleteRoute)
    delSubmitButton.grid(row=1, column=1, columnspan=2)

    del_root.bind('<Return>', deleteRoute)

# Goal Graph
def setGoal():
    def setGraph(event):
        if entry_v0.get() != '' and entry_v0.get() != '' and entry_v1.get() != '' \
                and entry_v2.get() != '' and entry_v3.get() != '' and entry_v4.get() != '' \
                and entry_v5.get() != '' and entry_v6.get() != '' and entry_v7.get() != '' \
                and entry_v8.get() != '' and entry_v9.get() != '' and entry_6.get() != '' \
                and entry_6.get() != '' and entry_7.get() != '' and entry_8.get() != '' \
                and entry_9.get() != '' and entry_10.get() != '' and entry_11.get() != '' \
                and entry_12.get() != '' and entry_13.get() != '':
            # Update spreadsheet
            data_sheet.cell(row=2, column=1).value = int(entry_v0.get())
            data_sheet.cell(row=3, column=1).value = int(entry_v1.get())
            data_sheet.cell(row=4, column=1).value = int(entry_v2.get())
            data_sheet.cell(row=5, column=1).value = int(entry_v3.get())
            data_sheet.cell(row=6, column=1).value = int(entry_v4.get())
            data_sheet.cell(row=7, column=1).value = int(entry_v5.get())
            data_sheet.cell(row=8, column=1).value = int(entry_v6.get())
            data_sheet.cell(row=9, column=1).value = int(entry_v7.get())
            data_sheet.cell(row=10, column=1).value = int(entry_v8.get())
            data_sheet.cell(row=11, column=1).value = int(entry_v9.get())

            # Ropes
            data_sheet.cell(row=2, column=2).value = int(entry_6.get())
            data_sheet.cell(row=3, column=2).value = int(entry_7.get())
            data_sheet.cell(row=4, column=2).value = int(entry_8.get())
            data_sheet.cell(row=5, column=2).value = int(entry_9.get())
            data_sheet.cell(row=6, column=2).value = int(entry_10.get())
            data_sheet.cell(row=7, column=2).value = int(entry_11.get())
            data_sheet.cell(row=8, column=2).value = int(entry_12.get())
            data_sheet.cell(row=9, column=2).value = int(entry_13.get())

            g_graph_root.destroy()
            route_book.save('Routes.xlsx')

        else:
            tkinter.messagebox.showerror(title=None, message='One or more options was not filled out, please try again')

    route_book = openpyxl.load_workbook('Routes.xlsx')
    data_sheet = route_book['Data']
    # Creates a window with entry boxes for every grade type
    g_graph_root = tkinter.Toplevel(root)

    main_label = Label(g_graph_root, text='Enter your preferred route curve below:')
    main_label.grid(row=0, column=1, columnspan=2)

    label_v0 = Label(g_graph_root, text='V0s')
    label_v0.grid(row=1, column=0)
    entry_v0 = Entry(g_graph_root)
    entry_v0.grid(row=1, column=1)

    label_v1 = Label(g_graph_root, text='V1s')
    label_v1.grid(row=2, column=0)
    entry_v1 = Entry(g_graph_root)
    entry_v1.grid(row=2, column=1)

    label_v2 = Label(g_graph_root, text='V2s')
    label_v2.grid(row=3, column=0)
    entry_v2 = Entry(g_graph_root)
    entry_v2.grid(row=3, column=1)

    label_v3 = Label(g_graph_root, text='V3s')
    label_v3.grid(row=4, column=0)
    entry_v3 = Entry(g_graph_root)
    entry_v3.grid(row=4, column=1)

    label_v4 = Label(g_graph_root, text='V4s')
    label_v4.grid(row=5, column=0)
    entry_v4 = Entry(g_graph_root)
    entry_v4.grid(row=5, column=1)

    label_v5 = Label(g_graph_root, text='V5s')
    label_v5.grid(row=6, column=0)
    entry_v5 = Entry(g_graph_root)
    entry_v5.grid(row=6, column=1)

    label_v6 = Label(g_graph_root, text='V6s')
    label_v6.grid(row=7, column=0)
    entry_v6 = Entry(g_graph_root)
    entry_v6.grid(row=7, column=1)

    label_v7 = Label(g_graph_root, text='V7s')
    label_v7.grid(row=8, column=0)
    entry_v7 = Entry(g_graph_root)
    entry_v7.grid(row=8, column=1)

    label_v8 = Label(g_graph_root, text='V8s')
    label_v8.grid(row=9, column=0)
    entry_v8 = Entry(g_graph_root)
    entry_v8.grid(row=9, column=1)

    label_v9 = Label(g_graph_root, text='V9s')
    label_v9.grid(row=10, column=0)
    entry_v9 = Entry(g_graph_root)
    entry_v9.grid(row=10, column=1)

    # Creates a window with entry boxes for every grade type
    label_6 = Label(g_graph_root, text='5.6s')
    label_6.grid(row=2, column=2)
    entry_6 = Entry(g_graph_root)
    entry_6.grid(row=2, column=3)

    label_7 = Label(g_graph_root, text='5.7s')
    label_7.grid(row=3, column=2)
    entry_7 = Entry(g_graph_root)
    entry_7.grid(row=3, column=3)

    label_8 = Label(g_graph_root, text='5.8s')
    label_8.grid(row=4, column=2)
    entry_8 = Entry(g_graph_root)
    entry_8.grid(row=4, column=3)

    label_9 = Label(g_graph_root, text='5.9s')
    label_9.grid(row=5, column=2)
    entry_9 = Entry(g_graph_root)
    entry_9.grid(row=5, column=3)

    label_10 = Label(g_graph_root, text='5.10s')
    label_10.grid(row=6, column=2)
    entry_10 = Entry(g_graph_root)
    entry_10.grid(row=6, column=3)

    label_11 = Label(g_graph_root, text='5.11s')
    label_11.grid(row=7, column=2)
    entry_11 = Entry(g_graph_root)
    entry_11.grid(row=7, column=3)

    label_12 = Label(g_graph_root, text='5.12s')
    label_12.grid(row=8, column=2)
    entry_12 = Entry(g_graph_root)
    entry_12.grid(row=8, column=3)

    label_13 = Label(g_graph_root, text='5.13s')
    label_13.grid(row=9, column=2)
    entry_13 = Entry(g_graph_root)
    entry_13.grid(row=9, column=3)

    get_graph = Button(g_graph_root, text='Set Graphs')
    get_graph.bind('<Button-1>', setGraph)
    get_graph.grid(row=11, column=1, columnspan=2, padx=50)

# Show current and goal graph
def graphRoutes(sheet):
    # Get data for current graph
    route_book = openpyxl.load_workbook('Routes.xlsx')
    entry_sheet = route_book[sheet]

    if sheet == 'Boulders':
        grade_list = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        x_axis = ['V0', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6', 'V7', 'V8', 'V9']
        replace = 'V'
        grade_minus = 0
        data_column = 1
    else:
        grade_list = [0, 0, 0, 0, 0, 0, 0, 0]
        x_axis = ['5.6', '5.7', '5.8', '5.9', '5.10', '5.11', '5.12', '5.13']
        replace = '5.'
        grade_minus = 6
        data_column = 2
    positions = np.arange(len(grade_list))

    for cell in range(2, entry_sheet.max_row + 1):
        temp_grade = entry_sheet.cell(row=cell, column=1).value

        grade_int = int(temp_grade.replace(replace, ''))
        # Fill the list with appropriate values
        grade_list[grade_int - grade_minus] += 1

    # Get data for goal graph
    goal_list = []
    data_sheet = route_book['Data']
    for cell in range(2, data_sheet.max_row + 1):
        temp_g_grade = data_sheet.cell(row=cell, column=data_column).value
        if temp_g_grade is not None:
            goal_list.append(temp_g_grade)
            route_book.save('Routes.xlsx')

    try:
        plt.bar(positions, grade_list, width=0.5, color='red')
        plt.bar(positions + 0.5, goal_list, width=0.5)
        plt.legend(['Current Spread', 'Goal Spread'])
        plt.title('Route Comparison')
        plt.xticks(positions + 0.25, x_axis)
        plt.ylabel('Amount')

        plt.show()
    except ValueError:
        tkinter.messagebox.showerror('Missing Data', 'You have not submitted your ideal route curve. '
                                                     'Please go to settings')

# Fill data sheet
def fillData():
    def fillSet(col):
        entry = []
        for cell in range(2, data_sheet.max_row + 1):
            cell_data = data_sheet.cell(row=cell, column=col).value

            if cell_data is not None:
                entry.append(cell_data)

        return entry

    def fillSheet(entry, col):
        for cell in range(len(entry)):
            data_sheet.cell(row=cell+2, column=col).value = entry[cell]

    def stripSpaces(entry):
        list = entry.split(', ')
        for item in list:
            item.replace(' ', '')

        return list

    def overwriteData():
        setter_list = stripSpaces(setter_entry.get())
        color_list = stripSpaces(color_entry.get())
        wall_list = stripSpaces(wall_entry.get())
        rwall_list = stripSpaces(rwall_entry.get())

        # Fill spreadsheet
        # Setters
        for s in range(len(setter_list)):
            data_sheet.cell(row=s+2, column=3).value = setter_list[s]

        # Colors
        for c in range(len(color_list)):
            data_sheet.cell(row=c+2, column=4).value = color_list[c]

        # Walls
        for w in range(len(wall_list)):
            data_sheet.cell(row=w+2, column=5).value = wall_list[w]

        # Rope Walls
        for r in range(len(rwall_list)):
            data_sheet.cell(row=r+2, column=6).value = rwall_list[r]

        route_book.save('Routes.xlsx')
        updateOptions()
        data_root.destroy()

    def submitData():
        # Create sets out of current options
        setter_set = set(fillSet(3))
        color_set = set(fillSet(4))
        wall_set = set(fillSet(5))
        rwall_set = set(fillSet(6))

        setter_list = stripSpaces(setter_entry.get())
        setter_set.update(setter_list)

        color_list = stripSpaces(color_entry.get())
        color_set.update(color_list)

        wall_list = stripSpaces(wall_entry.get())
        wall_set.update(wall_list)

        rwall_list = stripSpaces(rwall_entry.get())
        rwall_set.update(rwall_list)

        setter_list = list(setter_set)
        color_list = list(color_set)
        wall_list = list(wall_set)
        rwall_list = list(rwall_set)

        # Update spreadsheet
        fillSheet(setter_list, 3)
        fillSheet(color_list, 4)
        fillSheet(wall_list, 5)
        fillSheet(rwall_list, 6)

        route_book.save('Routes.xlsx')
        updateOptions()
        data_root.destroy()

    # Open excel document
    try:
        route_book = openpyxl.load_workbook('Routes.xlsx')
        data_sheet = route_book['Data']
    except:
        warning_label.destroy()
        # Create new workbook
        route_book = openpyxl.Workbook()

        # Create boulder sheet
        entry_sheet = route_book.active
        entry_sheet.title = 'Boulders'

        # Set up first row
        entry_sheet['A1'] = 'Grade'
        entry_sheet['B1'] = 'Wall'
        entry_sheet['C1'] = 'Setter'
        entry_sheet['D1'] = 'Color'
        entry_sheet.freeze_panes = 'A2'

        # Create route sheet
        route_book.create_sheet('Ropes')
        rope_sheet = route_book['Ropes']

        # Set up first row
        rope_sheet['A1'] = 'Grade'
        rope_sheet['B1'] = 'Wall'
        rope_sheet['C1'] = 'Setter'
        rope_sheet['D1'] = 'Color'
        rope_sheet.freeze_panes = 'A2'

        # Set up data sheet
        route_book.create_sheet('Data')
        data_sheet = route_book['Data']

        # Set up first row
        data_sheet['A1'] = 'Boulder'
        data_sheet['B1'] = 'Ropes'
        data_sheet['C1'] = 'Setters'
        data_sheet['D1'] = 'Colors'
        data_sheet['E1'] = 'Walls'
        data_sheet['F1'] = 'RWalls'
        data_sheet.freeze_panes = 'A2'

        route_book.save('Routes.xlsx')

    data_root = tkinter.Toplevel(root)
    data_root.geometry('400x300')

    data_label = Label(data_root, text='Fill out the fields below, separating values with commas.')
    data_label.grid(row=0, column=0, columnspan=2, sticky=NSEW, pady=5)

    setter_label = Label(data_root, text='Setters:')
    setter_label.grid(row=1, column=0, sticky=NSEW, pady=5)
    setter_entry = Entry(data_root)
    setter_entry.grid(row=1, column=1, sticky=NSEW, pady=5)

    color_label = Label(data_root, text='Colors:')
    color_label.grid(row=2, column=0, sticky=NSEW, pady=5)
    color_entry = Entry(data_root)
    color_entry.grid(row=2, column=1, sticky=NSEW, pady=5)

    wall_label = Label(data_root, text='Walls:')
    wall_label.grid(row=3, column=0, sticky=NSEW, pady=5)
    wall_entry = Entry(data_root)
    wall_entry.grid(row=3, column=1, sticky=NSEW, pady=5)

    rwall_label = Label(data_root, text='Rope Walls:')
    rwall_label.grid(row=4, column=0, sticky=NSEW, pady=5)
    rwall_entry = Entry(data_root)
    rwall_entry.grid(row=4, column=1, sticky=NSEW, pady=5)

    data_submit = Button(data_root, text='Submit', command=submitData)
    data_submit.grid(row=5, column=0, columnspan=2, sticky=NSEW, pady=5, padx=150)

    data_overwrite = Button(data_root, text='Overwrite', command=overwriteData)
    data_overwrite.grid(row=6, column=0, columnspan=2, sticky=NSEW, pady=5, padx=150)

    for num in range(0, 7):
        data_root.grid_columnconfigure(num, weight=1)
        data_root.grid_rowconfigure(num, weight=1)

def highlightSubmit(*args):
    global wallDrop
    global gradeDrop
    global colorDrop
    global setterDrop
    global rwallDrop
    global rGradeDrop

    if wallDrop.get() != 'Wall:' and colorDrop.get() != 'Color:' and setterDrop.get() != 'Setter:'\
        and gradeDrop.get() != 'Grade:':
        submitButton.configure(state=ACTIVE)

    if rwallDrop.get() != 'Wall:' and colorDrop.get() != 'Color:' and setterDrop.get() != 'Setter:'\
        and rGradeDrop.get() != 'Grade:':
        rSubmitButton.configure(state=ACTIVE)

# Initialize tkinter module
root = Tk()
root.title('onSet')
tabControl = ttk.Notebook(root)

boulderTab = ttk.Frame(tabControl)
tabControl.add(boulderTab, text='Boulders')

ropeTab = ttk.Frame(tabControl)
tabControl.add(ropeTab, text='Ropes')

settingTab = ttk.Frame(tabControl)
tabControl.add(settingTab, text='Settings')

tabControl.pack(expand=1, fill='both')

# Define route variables and initial values
rope_grade_options = ['Grade:', '5.6', '5.7', '5.8', '5.9', '5.10', '5.11', '5.12', '5.13']
grade_options = ['Grade:', 'V0', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6', 'V7', 'V8', 'V9']
rwall_options = ['Wall:']
wall_options = ['Wall:']
setter_options = ['Setter:']
color_options = ['Color:']
delete_list = {}

gradeDrop = StringVar()
gradeDrop.set(grade_options[0])

rGradeDrop = StringVar()
rGradeDrop.set(rope_grade_options[0])

rwallDrop = StringVar()
rwallDrop.set(rwall_options[0])

wallDrop = StringVar()
wallDrop.set(wall_options[0])

setterDrop = StringVar()
setterDrop.set(setter_options[0])

colorDrop = StringVar()
colorDrop.set(color_options[0])

# Track dropdown changes to enable submit button
gradeDrop.trace('w', highlightSubmit)
rGradeDrop.trace('w', highlightSubmit)
rwallDrop.trace('w', highlightSubmit)
wallDrop.trace('w', highlightSubmit)
setterDrop.trace('w', highlightSubmit)
colorDrop.trace('w', highlightSubmit)

# Boulder tab
submitButton = Button(boulderTab, text='Submit Route', command=lambda x='Boulders': submitRoute(x), state=DISABLED)
submitButton.grid(row=2, column=1, sticky=NSEW)

deleteButton = Button(boulderTab, text='Delete Routes', command=lambda x='Boulders': deleteWindow(x))
deleteButton.grid(row=2, column=2, sticky=NSEW)

graphButton = Button(boulderTab, text='Route Graphs', command=lambda x='Boulders': graphRoutes(x))
graphButton.grid(row=3, column=1, columnspan=2, sticky=NSEW)

gradeMenu = OptionMenu(boulderTab, gradeDrop, *grade_options)
gradeMenu.grid(row=1, column=0, sticky=NSEW)

wallMenu = OptionMenu(boulderTab, wallDrop, *wall_options)
wallMenu.grid(row=1, column=1, sticky=NSEW)

setterMenu = OptionMenu(boulderTab, setterDrop, *setter_options)
setterMenu.grid(row=1, column=2, sticky=NSEW)

colorMenu = OptionMenu(boulderTab, colorDrop, *color_options)
colorMenu.grid(row=1, column=3, sticky=NSEW)

# Rope tab
rSubmitButton = Button(ropeTab, text='Submit Route', command=lambda x='Ropes': submitRoute(x), state=DISABLED)
rSubmitButton.grid(row=2, column=1, sticky=NSEW)

rDeleteButton = Button(ropeTab, text='Delete Routes', command=lambda x='Ropes': deleteWindow(x))
rDeleteButton.grid(row=2, column=2, sticky=NSEW)

rGraphButton = Button(ropeTab, text='Route Graphs', command=lambda x='Ropes': graphRoutes(x))
rGraphButton.grid(row=3, column=1, columnspan=2, sticky=NSEW)

rope_grade_menu = OptionMenu(ropeTab, rGradeDrop, *rope_grade_options)
gradeMenu.grid(row=1, column=0, sticky=NSEW)

rope_wall_menu = OptionMenu(ropeTab, wallDrop, *rwall_options)
wallMenu.grid(row=1, column=1, sticky=NSEW)

rope_setter_menu = OptionMenu(ropeTab, setterDrop, *setter_options)
rope_setter_menu.grid(row=1, column=2, sticky=NSEW)

colorMenu = OptionMenu(ropeTab, colorDrop, *color_options)
colorMenu.grid(row=1, column=3, sticky=NSEW)

# Settings tab
dataButton = Button(settingTab, text='Choose Presets', command=fillData)
dataButton.grid(row=1, column=1, columnspan=2, sticky=NSEW)

goalButton = Button(settingTab, text='Set Ideal Route Curve', command=setGoal)
goalButton.grid(row=2, column=1, columnspan=2, sticky=NSEW)

# Calibrate grid
for num in range(0, 6):
    boulderTab.grid_columnconfigure(num, weight=1)
    boulderTab.grid_rowconfigure(num, weight=1)

    ropeTab.grid_columnconfigure(num, weight=1)
    ropeTab.grid_rowconfigure(num, weight=1)

    settingTab.grid_columnconfigure(num, weight=1)
    settingTab.grid_rowconfigure(num, weight=1)

def updateOptions():
    global warning_label
    global wall_options
    global setter_options
    global color_options
    global rwall_options

    # Open workbook
    route_book = openpyxl.load_workbook('Routes.xlsx')
    data_sheet = route_book['Data']

    # Reset option lists
    wall_options = ['Wall:']
    setter_options = ['Setter:']
    color_options = ['Color:']
    rwall_options = ['Wall:']

    for cell in range(2, data_sheet.max_row + 1):
        # Make variable out of cells
        setter_cell = data_sheet.cell(row=cell, column=3).value
        color_cell = data_sheet.cell(row=cell, column=4).value
        wall_cell = data_sheet.cell(row=cell, column=5).value
        rwall_cell = data_sheet.cell(row=cell, column=6).value

        # Iterate over variables, appending to appropriate lists
        if setter_cell is not None:
            setter_options.append(setter_cell)
        if color_cell is not None:
            color_options.append(color_cell)
        if wall_cell is not None:
            wall_options.append(wall_cell)
        if rwall_cell is not None:
            rwall_options.append(rwall_cell)

    # Place dropdowns in boulder menu
    gradeMenu = OptionMenu(boulderTab, gradeDrop, *grade_options)
    gradeMenu.grid(row=1, column=0, sticky=NSEW)

    wallMenu = OptionMenu(boulderTab, wallDrop, *wall_options)
    wallMenu.grid(row=1, column=1, sticky=NSEW)

    setterMenu = OptionMenu(boulderTab, setterDrop, *setter_options)
    setterMenu.grid(row=1, column=2, sticky=NSEW)

    colorMenu = OptionMenu(boulderTab, colorDrop, *color_options)
    colorMenu.grid(row=1, column=3, sticky=NSEW)

    # Place dropdowns in rope menu
    rope_grade_menu = OptionMenu(ropeTab, rGradeDrop, *rope_grade_options)
    rope_grade_menu.grid(row=1, column=0, sticky=NSEW)

    rope_wall_menu = OptionMenu(ropeTab, rwallDrop, *rwall_options)
    rope_wall_menu.grid(row=1, column=1, sticky=NSEW)

    rope_setter_menu = OptionMenu(ropeTab, setterDrop, *setter_options)
    rope_setter_menu.grid(row=1, column=2, sticky=NSEW)

    rope_color_menu = OptionMenu(ropeTab, colorDrop, *color_options)
    rope_color_menu.grid(row=1, column=3, sticky=NSEW)

# Attempt to open spreadsheet
try:
    route_book = openpyxl.load_workbook('Routes.xlsx')
    data_sheet = route_book['Data']
    updateOptions()
except FileNotFoundError:
    warning_label = Label(boulderTab, text='No spreadsheet detected, please open settings.', fg='red')
    warning_label.grid(row=0, column=0, columnspan=4)

root.mainloop()
