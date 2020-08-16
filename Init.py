import openpyxl
from tkinter import *
import tkinter.messagebox

# Create new workbook
route_book = openpyxl.Workbook()

# Create boulder sheet
boulder_sheet = route_book.active
boulder_sheet.title = 'Boulders'

# Set up first row
boulder_sheet['A1'] = 'Grade'
boulder_sheet['B1'] = 'Wall'
boulder_sheet['C1'] = 'Setter'
boulder_sheet['D1'] = 'Color'
boulder_sheet.freeze_panes = 'A2'

# Create route sheet
route_book.create_sheet('Ropes')
rope_sheet = route_book['Ropes']

# Set up first row
rope_sheet['A1'] = 'Grade'
rope_sheet['B1'] = 'Wall'
rope_sheet['C1'] = 'Setter'
rope_sheet['D1'] = 'Color'
rope_sheet.freeze_panes = 'A2'

# Set up data sheet
route_book.create_sheet('Data')
data_sheet = route_book['Data']

# Set up first row
data_sheet['A1'] = 'Boulder'
data_sheet['B1'] = 'Ropes'
data_sheet['C1'] = 'Setters'
data_sheet['D1'] = 'Colors'
data_sheet['E1'] = 'Walls'
data_sheet['F1'] = 'RWalls'
data_sheet.freeze_panes = 'A2'

# Fill data sheet
def fillData():
    def submitData():
        # Convert strings to lists
        setter_list = setter_entry.get().split(', ')
        color_list = color_entry.get().split(', ')
        wall_list = wall_entry.get().split(', ')
        rwall_list = rwall_entry.get().split(', ')

        # Fill spreadsheet
        # Setters
        for s in range(len(setter_list)):
            data_sheet.cell(row=s+2, column=3).value = setter_list[s]

        # Colors
        for c in range(len(color_list)):
            data_sheet.cell(row=c+2, column=4).value = color_list[c]

        # Walls
        for w in range(len(wall_list)):
            data_sheet.cell(row=w+2, column=5).value = wall_list[w]

        # Rope Walls
        for r in range(len(rwall_list)):
            data_sheet.cell(row=r+2, column=6).value = rwall_list[r]

        data_root.destroy()

    data_root = Tk()
    data_root.geometry('400x300')

    data_label = Label(data_root, text='Fill out the fields below, separating values with commas.')
    data_label.grid(row=0, column=0, columnspan=2, sticky=NSEW, pady=5)

    setter_label = Label(data_root, text='Setters:')
    setter_label.grid(row=1, column=0, sticky=NSEW, pady=5)
    setter_entry = Entry(data_root)
    setter_entry.grid(row=1, column=1, sticky=NSEW, pady=5)

    color_label = Label(data_root, text='Colors:')
    color_label.grid(row=2, column=0, sticky=NSEW, pady=5)
    color_entry = Entry(data_root)
    color_entry.grid(row=2, column=1, sticky=NSEW, pady=5)

    wall_label = Label(data_root, text='Walls:')
    wall_label.grid(row=3, column=0, sticky=NSEW, pady=5)
    wall_entry = Entry(data_root)
    wall_entry.grid(row=3, column=1, sticky=NSEW, pady=5)

    rwall_label = Label(data_root, text='Rope Walls:')
    rwall_label.grid(row=4, column=0, sticky=NSEW, pady=5)
    rwall_entry = Entry(data_root)
    rwall_entry.grid(row=4, column=1, sticky=NSEW, pady=5)

    data_submit = Button(data_root, text='Submit', command=submitData)
    data_submit.grid(row=5, column=0, columnspan=2, sticky=NSEW, pady=5, padx=150)

    data_root.grid_columnconfigure(0, weight=1)
    data_root.grid_columnconfigure(1, weight=1)
    data_root.grid_columnconfigure(2, weight=1)
    data_root.grid_columnconfigure(3, weight=1)

    data_root.grid_rowconfigure(0, weight=1)
    data_root.grid_rowconfigure(1, weight=1)
    data_root.grid_rowconfigure(2, weight=1)
    data_root.grid_rowconfigure(3, weight=1)
    data_root.grid_rowconfigure(4, weight=1)
    data_root.grid_rowconfigure(5, weight=1)

    data_root.mainloop()
fillData()

# Get goal graph data
setGoal()
