# Imports
import matplotlib.pyplot as plt
import openpyxl
import numpy as np

# Open the excel document
routeBook = openpyxl.load_workbook('Routes.xlsx')
boulderSheet = routeBook['Boulders']
ropeSheet = routeBook['Ropes']

routeOrBoulder = str(input('Do you want the route or boulder graph? '))
if routeOrBoulder.lower() == 'boulder' or routeOrBoulder.lower() == 'boulders':
    grade_list = [0,0,0,0,0,0,0,0,0,0]
    x_axis = []
    positions = np.arange(len(grade_list))
    for i in range(0,10):
        x_axis.append('V' + str(i))
    for cell in range(2,boulderSheet.max_row + 1):
        temp_grade = boulderSheet.cell(column=1, row=cell).value
        grade_int = int(temp_grade.replace('V', ''))
        # Fill the list with appropriate values
        if grade_int == 0:
            grade_list[0] += 1
        elif grade_int == 1:
            grade_list[1] += 1
        elif grade_int == 2:
            grade_list[2] += 1
        elif grade_int == 3:
            grade_list[3] += 1
        elif grade_int == 4:
            grade_list[4] += 1
        elif grade_int == 5:
            grade_list[5] += 1
        elif grade_int == 6:
            grade_list[6] += 1
        elif grade_int == 7:
            grade_list[7] += 1
        elif grade_int == 8:
            grade_list[8] += 1
        elif grade_int == 9:
            grade_list[9] += 1

elif routeOrBoulder.lower() == 'rope' or routeOrBoulder.lower() == 'route' or routeOrBoulder == 'ropes':
    grade_list = [0,0,0,0,0,0,0,0]
    x_axis = []
    positions = np.arange(len(grade_list))
    for i in range(6,14):
        x_axis.append('5.' + str(i))
    for cell in range(2,ropeSheet.max_row + 1):
        temp_grade = ropeSheet.cell(column=1, row=cell).value
        grade_int = int(temp_grade.replace('5.', ''))
        # Fill the list with appropriate values
        if grade_int == 6:
            grade_list[0] += 1
        elif grade_int == 7:
            grade_list[1] += 1
        elif grade_int == 8:
            grade_list[2] += 1
        elif grade_int == 9:
            grade_list[3] += 1
        elif grade_int == 10:
            grade_list[4] += 1
        elif grade_int == 11:
            grade_list[5] += 1
        elif grade_int == 12:
            grade_list[6] += 1
        elif grade_int == 13:
            grade_list[7] += 1

# Show resulting graph
plt.bar(positions, grade_list)
plt.xticks(positions, x_axis)
plt.xlabel('Grades')
plt.ylabel('Amount')
plt.show()
print(x_axis)
routeBook.save('Routes.xlsx')